import openpyxl,csv,re,os,io
from operator import itemgetter

def readBusinessxlsx(Sfileame):
    print("Reading xlsx:",Sfileame)
    wb = openpyxl.load_workbook(filename=Sfileame)
    ws = wb.worksheets[0]
    return [[cell.value for cell in row] for row in ws.iter_rows()]

def readPriorxlsx(Sfileame):
    print("Reading xlsx:",Sfileame)
    wb = openpyxl.load_workbook(filename=Sfileame)
    ws = wb.worksheets[0]
    return [[cell.value for cell in row] for row in ws.iter_rows()]

def readUSL(Sfilename):
    return [line.split(',') for line in io.open(Sfilename,'r', encoding='utf-16-le')]

def getIdsHash(Sfileame):
    dataArr=readBusinessxlsx(Sfileame)
    IDshash={heading.upper():[] for index,heading in enumerate(dataArr[0]) if heading}
    indecesForIds={index:heading.upper() for index,heading in enumerate(dataArr[0]) if heading}
    #print(indecesForIds,IDshash)
    for row in dataArr[1:]:
        for index in indecesForIds.keys():
            if row[index] and row[index] not in IDshash[indecesForIds[index]]:
                IDshash[indecesForIds[index]].append(row[index])
    return IDshash


def open_csv_with_exception_hanlding(Sfileame):
    f = open(Sfileame)
    #detect_null_character_in_csv
    if '\0' in open(Sfileame).read():
        return csv.reader(x.replace('\0', '') for x in f)
    else:
        return csv.reader(f)

def handling_for_null_character_case(row):
    if len(row)==1 and re.search('\t',row[0]):
            row=row[0].split('\t')
    return row

def readcsv(Sfileame):
    print("Reading csv:",Sfileame)
    csv_f = open_csv_with_exception_hanlding(Sfileame)
    return [handling_for_null_character_case(row) for row in csv_f]

def readxlsx(Sfileame):
    print("Reading xlsx:",Sfileame)
    wb = openpyxl.load_workbook(filename=Sfileame,data_only=True)
    ws = wb.worksheets[0]
    return [[cell.value for cell in row] for row in ws.iter_rows()]

def readxlsm(Sfilename):
    #mainly handling for PaymentTech Discover Files
    print("Reading xlsm:",Sfilename)
    wb = openpyxl.load_workbook(filename=Sfilename,data_only=True)
    dataarr=[]
    for ws in wb.worksheets[1:]:
        for row in [[cell.value for cell in row] for row in ws.iter_rows()]:
            dataarr.append(row)
    return dataarr

def writeOutput(Dataarr,Ofilename):
    wb=openpyxl.Workbook()
    ws = wb.active

    for row in Dataarr:
        ws.append([val.replace(',','') if type(val) is str else val for val in row])

    wb.save(Ofilename)

    print(Ofilename,' File Loaded and Saved!!')


def getDirName(directory,year,month):
    directory=directory.replace("YYYY",str(year))
    directory=directory.replace('MM.YY',str(month)+'.'+str(year)[2:])
    return directory

def getFileName(filename,year,month):
    filename=filename.replace("YYYY",str(year))
    filename=filename.replace('MM.YY',str(month)+'.'+str(year)[2:])
    return filename



cleansing_of_given_processor={
    "DE AMEX REPORTS":None,
    "EU AMEX REPORTS":None,
    "US AMEX REPORTS":None,
    "FR AMEX REPORTS":None,
    "AU AMEX REPORTS":None,
    "UK AMEX REPORTS":None,
    "CA AMEX REPORTS":None,
    "BAMS":None,
    "BILLDESK":None,
    "DISCOVER":None,
    "HDFC":None,
    "NETEASE":None,
    "PAYMENTECH":None
}

cleansing_of_given_processor={
    "DE AMEX REPORTS":None,
    "EU AMEX REPORTS":None,
    "US AMEX REPORTS":None,
    "FR AMEX REPORTS":None,
    "AU AMEX REPORTS":None,
    "UK AMEX REPORTS":None,
    "CA AMEX REPORTS":None,
    "BAMS":None,
    "BILLDESK":None,
    "DISCOVER":None,
    "HDFC":None,
    "NETEASE":None,
    "PAYMENTECH":None
}

decodingValues={

            'EINREICHUNGEN':'SUBMISSIONS',
            '':'ADJUSTMENTS',
            'RÜCKBELASTUNGEN':'CHARGEBACKS',
            'ÜBERSICHT DER GESAMTSUMMEN':'SUMMARY TOTALS'
}

data_to_pick_for_output={

    "FR AMEX REPORTS":{'columns':[],'heading':[]},
    "AU AMEX REPORTS":{'columns':[],'heading':[]},
    "DE AMEX REPORTS":{'columns':[],'heading':[]},

    "EU AMEX REPORTS":{
        'heading':["Submitting Merchant #","Activity Type","Submission Amount","Flip Discount Amount","Flip Fees Incentives","Settlement Date"],
        'SUBMISSIONS':{
            'columns':[2,21,8,11,12,0],
            'id':2,
            'sdateIndex':0,
            'sdateformat':'%d/%m/%Y'
        },
        'ADJUSTMENTS':{
            'columns':[0,21,4,7,9,10],
            'id':0,
            'sdateIndex':10,
            'sdateformat':'%d/%m/%Y'
        },
        'CHARGEBACKS':{
            'columns':[0,22,21,6,7,9],
            'id':0,
            'sdateIndex':9,
            'sdateformat':'%d/%m/%Y'
        }
    },

    "UK AMEX REPORTS":{
        'heading':["Submitting Merchant #","Activity Type","Submission Amount","Flip Discount Amount","Flip Fees Incentives","Settlement Date"],
        'SUBMISSIONS':{
            'columns':[2,21,6,11,12,0],
            'id':2,
            'sdateIndex':0,
            'sdateformat':'%d/%m/%Y'
        }
    },




    "US AMEX REPORTS":{
        'heading':["Submitting Merchant #","Activity Type","Submission Amount","Flip Discount Amount","Flip Fees Incentives","Settlement Date"],
        'SUBMISSIONS':{
            'columns':[2,12,9,3,4,0],
            'id':2,
            'sdateIndex':0,
            'sdateformat':'%m/%d/%Y'
        },
        'ADJUSTMENTS':{
            'columns':[0,14,4,7,8,10],
            'id':0,
            'sdateIndex':10,
            'sdateformat':'%m/%d/%Y'
        },
        'CHARGEBACKS':{
            'columns':[0,13,2,4,5,7],
            'id':0,
            'sdateIndex':7,
            'sdateformat':'%m/%d/%Y'
        }
    },

    "CA AMEX REPORTS":{
        'heading':["Submitting Merchant #","Activity Type","Submission Amount","Flip Discount Amount","Flip Fees Incentives","Settlement Date"],
        'SUBMISSIONS':{
            'columns':[2,12,9,3,4,0],
            'id':2,
            'sdateIndex':0,
            'sdateformat':'%d/%m/%Y'
        },
        'ADJUSTMENTS':{
            'columns':[0,14,4,7,8,10],
            'id':0,
            'sdateIndex':10,
            'sdateformat':'%d/%m/%Y'
        }
    },
    "BAMS":{'columns':[1,2,10,11,12,13],
            'heading':["Funded Date","Merchant ID","Processed I/C Charges","Processed Service Charges","Processed Fees","Processed Chargebacks/Reversals"],
            'id':2
            },
    "BILLDESK":{'columns':[1,2,6,7],
                'heading':["Processing Division","Transaction Date","Chargebacks","Fees"],
                'id':1
                },
    "DISCOVER":{'columns':[8,7,0,4,6],
                'heading':["Division Name","Discover","Activity Ending Date","Chargebacks & Adjustments","Discount Activity"],
                'id':7
            },
    "HDFC":{'columns':[1,2,6,7],
            'heading':["Processing Division","Transaction Date","Chargebacks","Fees"],
            'id':1
            },
    "NETEASE":{'columns':[1,2,6,7],
               'heading':["Processing Division","Transaction Date","Chargebacks","Fees"],
               'id':1
               },
    "PAYMENTECH":{'columns':[4,5,10,11,12],
                  'heading':["Entity ID","Merchant Activity Date","Chargeback and ECP Return Adjustments","Interchange & Assessment Fees","Paymentech Fees"],
                  'id':4
                  }

}

TabNames={

    "EU AMEX REPORTS":"AMEX",
    "US AMEX REPORTS":"AMEX",
    "UK AMEX REPORTS":"AMEX",
    "CA AMEX REPORTS":"AMEX",
    "BAMS":"BAMS",
    "BILLDESK":"BILLDESK",
    "DISCOVER":"DISCOVER",
    "HDFC":"HDFC",
    "NETEASE":"NETEASE",
    "PAYMENTECH":"PAYMENTECH"
    #"FR AMEX REPORTS":"AMEX",
    #"AU AMEX REPORTS":"AMEX",
    #"DE AMEX REPORTS":"AMEX"
}



filetypesinfolders={

    "EU AMEX REPORTS":['csv'],
    "US AMEX REPORTS":['csv'],
    "UK AMEX REPORTS":['csv'],
    "CA AMEX REPORTS":['csv'],
    "BAMS":['csv'],
    "BILLDESK":['xlsx','xls'],
    "DISCOVER":['xlsm','xls'],
    "HDFC":['xlsx','xls'],
    "NETEASE":['csv'],
    "PAYMENTECH":['xlsx','xls']
    #"DE AMEX REPORTS":['csv'],
    #"FR AMEX REPORTS":['csv'],
    #"AU AMEX REPORTS":['csv'],

}


scriptName="FiltrationProcess.py"

directory=r"C:\Users\sssumit\Project\FUSE REPORT\TEST PYTHON\YYYY\MM.YY\Entry Support\Processor Reports"
USLFilename=r"C:\Users\sssumit\Project\FUSE REPORT\TEST PYTHON\YYYY\MM.YY\Entry Support\Processor Reports\Common\USL.txt"
PriorTransitName=r"C:\Users\sssumit\Project\FUSE REPORT\TEST PYTHON\YYYY\MM.YY\Entry Support\Processor Reports\Common\PriorTransitAmts.xlsx"
Businessfilename=r"C:\Users\sssumit\Project\FUSE REPORT\TEST PYTHON\YYYY\MM.YY\Entry Support\Processor Reports\Common\Business_List.xlsx"

TempDir=os.getcwd()+r"\OUTPUT"
if not os.path.exists(TempDir):os.makedirs(TempDir)
MasterFile=TempDir+ r"\MasterOutPut.xlsx"


readfile_of_given_type={'csv':None, 'xlsx':None, 'xlsm':None}

